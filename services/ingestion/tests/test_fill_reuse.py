"""AI fill-reuse — unchanged source content carries forward AI fills, no API call.

Measured finding (2026-06-12): live ``ai_map`` category fills are not byte-stable across
re-ingests — 55 ATC-less SL records re-versioned over two re-runs — and inversely a
key-less re-run would re-version filled records back to fill-less. Both break re-ingest
idempotency. Fill-reuse closes both: when the deterministic pre-fill content of a row is
unchanged from the latest frozen version, we adopt that version's content verbatim
(including its AI fills) with NO model call, so freeze reproduces the identical
``record_hash`` and the repository dedupes. ``--refill`` is the deliberate exception.

All offline: SQLite mirror, fake/absent anthropic client, no network.
"""

from __future__ import annotations

import sys
import types


from tarifhub_ingest.audit.audit_logger import AuditLogger
from tarifhub_ingest.config import get_settings
from tarifhub_ingest.ingestion.fill_reuse import strip_to_prefill
from tarifhub_ingest.ingestion.pipeline import run_pipeline
from tarifhub_ingest.ingestion.source_loader import SourceSpec
from tarifhub_ingest.mappers.tariff_mapper import map_raw
from tarifhub_ingest.models.tariff_model import Designation, TariffRecord, TariffSystem
from tarifhub_ingest.storage.db import Database
from tarifhub_ingest.storage.tariff_repository import TariffRepository
from tarifhub_ingest.versioning.freeze_record import compute_record_hash, freeze


# --------------------------------------------------------------------------- #
# Fakes: an anthropic module whose construction raises (zero-API-call proof)
# --------------------------------------------------------------------------- #


class _ExplodingAnthropic:
    """Stand-in for ``anthropic.Anthropic`` that raises if EVER constructed."""

    def __init__(self, *args, **kwargs):  # noqa: D401
        raise AssertionError("anthropic client must NOT be constructed on the reuse path")


def _install_exploding_anthropic(monkeypatch) -> None:
    module = types.ModuleType("anthropic")
    module.Anthropic = _ExplodingAnthropic  # type: ignore[attr-defined]
    monkeypatch.setitem(sys.modules, "anthropic", module)


def _xlsx_one_row(tmp_path, *, code: str, de: str) -> "object":
    """A minimal EAL-shaped XLSX with one DE-only row (so the AI seam has a category gap)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    de_sheet = wb.create_sheet("Deutsch")
    de_sheet.append(["Fachbereiche"])
    de_sheet.append(["Kapitel", "Pos.-Nr.", "TP", "Bezeichnung"])  # no discipline col -> category None
    de_sheet.append(["k", code, 76.5, de])
    tmp_path.mkdir(parents=True, exist_ok=True)
    path = tmp_path / "analysenliste_2026-01-01.xlsx"
    wb.save(path)
    return path


def _wire(tmp_path, name: str = "reuse.db"):
    db = Database.from_url(f"sqlite:///{tmp_path / name}")
    conn = db.connect()
    db.init_schema(conn)
    return db, conn, TariffRepository(conn, db), AuditLogger(conn, db)


_EAL_URL = "https://www.bag.admin.ch/de/analysenliste-al"


def _filled_prev_version(repo: TariffRepository, *, path, code: str) -> TariffRecord:
    """Seed a frozen 'previous version' carrying an AI category fill (ai_fields=[category]).

    Built from the SAME parsed XLSX row the pipeline will see (so the pre-fill content is
    byte-identical to the pipeline's candidate), then the AI category fill on top — exactly
    what a prior live run would have frozen.
    """
    from tarifhub_ingest.adapters import bag_eal
    from tarifhub_ingest.confidence.scorer import score

    raw = next(r for r in bag_eal.parse(path) if r["tariff_code"] == code)
    candidate = map_raw(raw, system=TariffSystem.EAL, source_url=_EAL_URL)
    filled = candidate.model_copy(update={
        "category": "Spezialnahrung",
        "metadata": {
            **candidate.metadata,
            "ai_assisted": True,
            "ai_model": "claude-opus-4-8",
            "ai_status": "ok",
            "ai_fields": ["category"],
        },
    })
    # Score + flag exactly as the pipeline does before freezing, so the stored row is a
    # faithful prior version (harmonization_confidence / requires_review are hashed).
    confidence = score(filled)
    filled = filled.model_copy(update={
        "harmonization_confidence": confidence,
        "requires_review": confidence < get_settings().review_threshold,
    })
    frozen = freeze(filled)
    assert repo.add(frozen) is True
    return frozen


# --------------------------------------------------------------------------- #
# strip / compare helper — byte-exact round-trip
# --------------------------------------------------------------------------- #


def test_strip_to_prefill_reverses_a_category_fill_byte_exactly():
    """A stored record whose ai_fields=[category] strips to exactly the candidate.

    Compared via compute_record_hash both ways — the equality the reuse decision rests on.
    """
    candidate = map_raw({"tariff_code": "1", "designation_de": "Milupa", "tax_points": "76.5"},
                        system=TariffSystem.EAL)
    stored = candidate.model_copy(update={
        "category": "Spezialnahrung",
        "metadata": {
            **candidate.metadata,
            "ai_assisted": True, "ai_model": "claude-opus-4-8",
            "ai_status": "ok", "ai_fields": ["category"],
        },
    })
    stripped = strip_to_prefill(stored)
    assert compute_record_hash(stripped) == compute_record_hash(candidate)


def test_strip_to_prefill_reverses_designation_fills():
    """ai_fields=[designation_fr, designation_it] strip back to None inside Designation."""
    candidate = map_raw({"tariff_code": "1", "designation_de": "X", "tax_points": "76.5"},
                        system=TariffSystem.EAL)
    stored = candidate.model_copy(update={
        "designation": Designation(de="X", fr="trad-fr", it="trad-it"),
        "metadata": {
            **candidate.metadata,
            "ai_assisted": True, "ai_model": "claude-opus-4-8",
            "ai_status": "ok", "ai_fields": ["designation_fr", "designation_it"],
        },
    })
    stripped = strip_to_prefill(stored)
    assert compute_record_hash(stripped) == compute_record_hash(candidate)


def test_strip_preserves_non_ai_metadata():
    """Stripping keeps source/mapper metadata, only resetting AI provenance keys."""
    candidate = map_raw({"tariff_code": "1", "designation_de": "X", "price_chf": "1.00",
                         "metadata": {"dossier_number": "16577"}}, system=TariffSystem.SL)
    stored = candidate.model_copy(update={
        "category": "Cat",
        "metadata": {**candidate.metadata, "ai_assisted": True, "ai_model": "m",
                     "ai_status": "ok", "ai_fields": ["category"]},
    })
    stripped = strip_to_prefill(stored)
    assert stripped.metadata["dossier_number"] == "16577"
    assert stripped.metadata["ai_assisted"] is False
    assert "ai_model" not in stripped.metadata
    assert "ai_fields" not in stripped.metadata


# --------------------------------------------------------------------------- #
# Pipeline reuse path
# --------------------------------------------------------------------------- #


def _spec(path):
    return SourceSpec(system=TariffSystem.EAL, kind="bag_eal", path=path, source_url=_EAL_URL)


def test_reuse_makes_zero_api_calls(tmp_path, monkeypatch):
    """Reuse path constructs NO anthropic client even with a (dummy) key set."""
    _install_exploding_anthropic(monkeypatch)
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-dummy")
    db, conn, repo, audit = _wire(tmp_path)
    path = _xlsx_one_row(tmp_path, code="1000", de="Milupa OS")
    _filled_prev_version(repo, path=path, code="1000")

    report = run_pipeline([_spec(path)], repo, audit, settings=get_settings())

    # Unchanged content -> adopted verbatim -> identical hash -> deduped. No client built.
    assert report.frozen == 0
    assert report.skipped_existing == 1
    conn.close()


def test_keyless_reuse_still_skips(tmp_path, monkeypatch):
    """Inverse-bug regression: no key + filled history + same raw still SKIPS (no re-version)."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)
    path = _xlsx_one_row(tmp_path, code="1000", de="Milupa OS")
    _filled_prev_version(repo, path=path, code="1000")

    report = run_pipeline([_spec(path)], repo, audit, settings=get_settings())

    assert report.frozen == 0
    assert report.skipped_existing == 1
    # Still exactly one version: the fill carried forward, no fill-less re-version.
    rows = conn.execute("SELECT version FROM tariff WHERE tariff_code='1000'").fetchall()
    assert [r["version"] for r in rows] == [1]
    conn.close()


def test_changed_content_takes_normal_path(tmp_path, monkeypatch):
    """Changed source content -> normal path (offline fallback fills nothing, new version)."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)
    seed_path = _xlsx_one_row(tmp_path, code="1000", de="Milupa OS")
    _filled_prev_version(repo, path=seed_path, code="1000")

    # Different German designation -> different pre-fill content -> NOT reuse.
    changed = _xlsx_one_row(tmp_path / "changed", code="1000", de="Milupa GA")
    report = run_pipeline([_spec(changed)], repo, audit, settings=get_settings())

    assert report.frozen == 1
    rows = conn.execute("SELECT version FROM tariff WHERE tariff_code='1000' ORDER BY version").fetchall()
    assert [r["version"] for r in rows] == [1, 2]
    conn.close()


def test_refill_bypasses_reuse(tmp_path, monkeypatch):
    """refill=True bypasses reuse: with no key the fallback fills nothing -> deterministic.

    The previously-filled category is NOT carried forward; re-mapping a DE-only row with
    no key yields category=None, a different hash, hence a new version.
    """
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)
    path = _xlsx_one_row(tmp_path, code="1000", de="Milupa OS")
    _filled_prev_version(repo, path=path, code="1000")

    report = run_pipeline([_spec(path)], repo, audit, settings=get_settings(), refill=True)

    # refill re-maps fresh (no carry-forward); fill-less differs -> a new version.
    assert report.frozen == 1
    rows = conn.execute("SELECT version FROM tariff WHERE tariff_code='1000' ORDER BY version").fetchall()
    assert [r["version"] for r in rows] == [1, 2]
    conn.close()


def test_no_previous_version_takes_normal_path(tmp_path, monkeypatch):
    """No prior version -> ai_map as today (offline: fallback fills nothing, freezes v1)."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)

    path = _xlsx_one_row(tmp_path, code="2000", de="Neuer Eintrag")
    report = run_pipeline([_spec(path)], repo, audit, settings=get_settings())

    assert report.frozen == 1
    conn.close()


def test_audit_detail_carries_reuse_provenance(tmp_path, monkeypatch):
    """The audit detail records ai_fills_reused + reused_from_version (NOT record metadata)."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)
    path = _xlsx_one_row(tmp_path, code="1000", de="Milupa OS")
    prev = _filled_prev_version(repo, path=path, code="1000")

    run_pipeline([_spec(path)], repo, audit, settings=get_settings())

    rows = conn.execute(
        "SELECT detail FROM audit_log WHERE tariff_code='1000' ORDER BY id DESC LIMIT 1"
    ).fetchall()
    import json
    detail = json.loads(rows[0]["detail"])
    assert detail["ai_fills_reused"] is True
    assert detail["reused_from_version"] == prev.version

    # The frozen record's metadata must NOT carry reuse provenance (byte-stability).
    stored = repo.get("1000", TariffSystem.EAL)
    assert "ai_fills_reused" not in stored.metadata
    assert "reused_from_version" not in stored.metadata
    conn.close()


# --------------------------------------------------------------------------- #
# Reuse-edge pins (codex P2 blind spots)
# --------------------------------------------------------------------------- #


def _xlsx_with_category(tmp_path, *, code: str, de: str, category: str):
    """An EAL-shaped XLSX whose discipline column flags ``category`` (source-owned)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    de_sheet = wb.create_sheet("Deutsch")
    de_sheet.append(["Fachbereiche"])
    de_sheet.append(["Kapitel", "Pos.-Nr.", "TP", "Bezeichnung", category])
    de_sheet.append(["k", code, 76.5, de, "Ja"])  # 'Ja' -> category resolves to the column header
    tmp_path.mkdir(parents=True, exist_ok=True)
    path = tmp_path / "analysenliste_2026-01-01.xlsx"
    wb.save(path)
    return path


def test_source_now_provides_prior_fill_mints_source_owned_version(tmp_path, monkeypatch):
    """Source NOW carries the category the AI previously filled -> no reuse, no API call.

    The stored v1 strips its AI category back to None; the new candidate carries the value
    natively, so strip-compare does NOT match -> the reuse short-circuit is skipped. ai_map
    then runs, but the candidate has no fillable gap (category present), so the gap-gate
    makes ZERO API calls (the exploding-anthropic guard would raise if a client were built).
    The result is a legitimate new version whose category is source-owned and whose
    provenance is ai_assisted=False (not the prior AI fill).
    """
    _install_exploding_anthropic(monkeypatch)
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-dummy")
    db, conn, repo, audit = _wire(tmp_path)

    # v1: the DE-only row whose category the AI filled to "Chemie".
    from tarifhub_ingest.adapters import bag_eal
    from tarifhub_ingest.confidence.scorer import score

    seed_path = _xlsx_one_row(tmp_path, code="1000", de="Vitamin D")
    raw_seed = next(r for r in bag_eal.parse(seed_path) if r["tariff_code"] == "1000")
    candidate_seed = map_raw(raw_seed, system=TariffSystem.EAL, source_url=_EAL_URL)
    filled = candidate_seed.model_copy(update={
        "category": "Chemie",
        "metadata": {
            **candidate_seed.metadata,
            "ai_assisted": True, "ai_model": "claude-opus-4-8",
            "ai_status": "ok", "ai_fields": ["category"],
        },
    })
    c = score(filled)
    filled = filled.model_copy(update={
        "harmonization_confidence": c,
        "requires_review": c < get_settings().review_threshold,
    })
    assert repo.add(freeze(filled)) is True

    # Re-ingest: the SAME row, but the source now owns the category "Chemie" natively.
    source_path = _xlsx_with_category(tmp_path / "src", code="1000", de="Vitamin D", category="Chemie")
    report = run_pipeline([_spec(source_path)], repo, audit, settings=get_settings())

    # Strip-compare did not match (stored strips to None, candidate has "Chemie"),
    # so a new version is minted; no API call was made (guard did not raise).
    assert report.frozen == 1
    rows = conn.execute(
        "SELECT version, category FROM tariff WHERE tariff_code='1000' ORDER BY version"
    ).fetchall()
    assert [r["version"] for r in rows] == [1, 2]

    latest = repo.get("1000", TariffSystem.EAL)
    assert latest.version == 2
    assert latest.category == "Chemie"  # source-owned, same value as the prior AI fill
    assert latest.metadata["ai_assisted"] is False  # provenance is source, not AI
    assert "ai_fields" not in latest.metadata
    conn.close()


def test_lossy_prior_reuse_full_pipeline_skips_and_keeps_raw_marker(tmp_path, monkeypatch):
    """A lossy-billing prior version re-ingested unchanged reuses (skip), raw_* intact.

    First pass freezes the lossy row (price 12.345 -> None + raw_price_chf + review).
    Re-running the SAME row must dedupe (skipped_existing==1, frozen==0), the audit detail
    carries the reuse provenance, and the stored record keeps its raw_* overflow marker.
    """
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)

    # An SL NDJSON row is awkward to hand-build; use a direct map_raw + the EAL pipeline
    # shape instead by writing a lossy TP into the EAL workbook (4dp scale, 1.23456 lossy).
    path = _xlsx_one_row(tmp_path, code="3000", de="Lossy probe")
    # Overwrite the TP cell with a lossy 5-dp value.
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["Deutsch"]
    ws.cell(row=3, column=3, value="1.23456")  # TP column, lossy at 4 dp
    wb.save(path)

    first = run_pipeline([_spec(path)], repo, audit, settings=get_settings())
    assert first.frozen == 1
    stored = repo.get("3000", TariffSystem.EAL)
    assert stored.tax_points is None
    assert stored.metadata["raw_tax_points"] == "1.23456"
    assert stored.requires_review is True

    second = run_pipeline([_spec(path)], repo, audit, settings=get_settings())
    assert second.frozen == 0
    assert second.skipped_existing == 1

    import json
    detail = json.loads(
        conn.execute(
            "SELECT detail FROM audit_log WHERE tariff_code='3000' ORDER BY id DESC LIMIT 1"
        ).fetchone()["detail"]
    )
    assert detail["ai_fills_reused"] is True
    assert detail["reused_from_version"] == 1

    # raw_* marker survives the reuse round-trip on the stored record.
    still = repo.get("3000", TariffSystem.EAL)
    assert still.metadata["raw_tax_points"] == "1.23456"
    assert still.tax_points is None
    rows = conn.execute("SELECT version FROM tariff WHERE tariff_code='3000'").fetchall()
    assert [r["version"] for r in rows] == [1]  # no re-version
    conn.close()


def test_multi_version_chain_reuses_against_latest(tmp_path, monkeypatch):
    """Reuse compares against the LATEST frozen version (v2), not v1.

    Seed v1 (no fills) then v2 (AI-filled, the current latest). Re-ingesting the unchanged
    raw must strip-compare against v2, adopt its fills, dedupe (skipped_existing==1), and
    record reused_from_version==2 — no v3 is minted.
    """
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db, conn, repo, audit = _wire(tmp_path)
    path = _xlsx_one_row(tmp_path, code="1000", de="Milupa OS")

    from tarifhub_ingest.adapters import bag_eal
    from tarifhub_ingest.confidence.scorer import score

    raw = next(r for r in bag_eal.parse(path) if r["tariff_code"] == "1000")
    candidate = map_raw(raw, system=TariffSystem.EAL, source_url=_EAL_URL)

    # v1: a fill-less prior version (e.g. an earlier run before the AI seam ran).
    v1 = candidate.model_copy(update={
        "harmonization_confidence": score(candidate),
        "requires_review": score(candidate) < get_settings().review_threshold,
    })
    assert repo.add(freeze(v1)) is True

    # v2: the AI-filled version — distinct content (category) -> a real new version.
    filled = candidate.model_copy(update={
        "category": "Spezialnahrung",
        "metadata": {
            **candidate.metadata,
            "ai_assisted": True, "ai_model": "claude-opus-4-8",
            "ai_status": "ok", "ai_fields": ["category"],
        },
    })
    cf = score(filled)
    filled = filled.model_copy(update={
        "harmonization_confidence": cf,
        "requires_review": cf < get_settings().review_threshold,
    })
    assert repo.add(freeze(filled)) is True
    assert repo.get("1000", TariffSystem.EAL).version == 2

    report = run_pipeline([_spec(path)], repo, audit, settings=get_settings())

    # Reuse adopts v2's fills -> dedupe, no v3.
    assert report.frozen == 0
    assert report.skipped_existing == 1
    rows = conn.execute("SELECT version FROM tariff WHERE tariff_code='1000' ORDER BY version").fetchall()
    assert [r["version"] for r in rows] == [1, 2]

    import json
    detail = json.loads(
        conn.execute(
            "SELECT detail FROM audit_log WHERE tariff_code='1000' ORDER BY id DESC LIMIT 1"
        ).fetchone()["detail"]
    )
    assert detail["ai_fills_reused"] is True
    assert detail["reused_from_version"] == 2
    conn.close()
