"""Build the committed BAG EAL test fixture FROM the real (gitignored) workbook.

Deterministic given the raw file: takes the first 25 parallel data rows across all
three language sheets, preserves the banner + header layout, then deliberately
DOCTORS three rows (in the FIXTURE ONLY) so the test suite exercises edge paths that
the 2026-01-01 edition happens not to contain:

  * an FR+IT-empty row    -> mapper must yield ``designation.fr/it = None``
  * a TP = 'nach Aufwand' -> mapper must yield ``tax_points = None`` (non-billing)
  * an all-disciplines-cleared row -> adapter must yield ``category = None``

Run from the service root:

    uv run python tests/fixtures/eal/make_fixture.py

The real file is never committed (``data/raw/`` is gitignored); the small fixture
and this script are. Re-running with the same raw file produces an identical fixture.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

# Repo-root raw file -> service-local fixture.
# .../services/ingestion/tests/fixtures/eal/make_fixture.py
_SERVICE_ROOT = Path(__file__).resolve().parents[3]
_REPO_ROOT = _SERVICE_ROOT.parents[1]
_RAW = _REPO_ROOT / "data" / "raw" / "eal" / "analysenliste_2026-01-01.xlsx"
_FIXTURE = Path(__file__).resolve().with_name("analysenliste_2026-01-01_fixture.xlsx")

_SHEETS = ("Deutsch", "Français", "Italiano")
_POS_HEADERS = {"Deutsch": "Pos.-Nr.", "Français": "No. Pos.", "Italiano": "No. pos."}
_DATA_ROWS = 25

# Which of the first _DATA_ROWS rows (0-indexed within the data slice) to doctor,
# and how. Position numbers are preserved so the three sheets stay parallel.
_FR_IT_EMPTY_ROW = 5  # blank the FR + IT designation cells (fixture only)
_TP_TEXT_ROW = 10  # set TP to the string 'nach Aufwand' (fixture only)
_NO_DISCIPLINE_ROW = 15  # clear every discipline flag column (fixture only)

# Discipline flag columns (German labels) cleared for the no-category edge row.
_DISCIPLINE_COLUMNS = (
    "Chemie",
    "Hämatologie",
    "Immunologie",
    "Genetik",
    "Mikrobiologie",
    "Spezialanalyse",
    "Basisanalyse",
)


def _find_header_index(rows: list[tuple], pos_header: str) -> int:
    for i, row in enumerate(rows[:10]):
        if len(row) > 1 and row[1] is not None and str(row[1]).strip() == pos_header:
            return i
    raise ValueError(f"header row with {pos_header!r} not found")


def build() -> Path:
    if not _RAW.exists():
        raise FileNotFoundError(
            f"raw EAL file not found at {_RAW}; download it before building the fixture"
        )

    source = load_workbook(filename=str(_RAW), read_only=True, data_only=True)
    out = Workbook()
    out.remove(out.active)  # drop the default empty sheet

    de_header_cols: dict[str, int] = {}

    for name in _SHEETS:
        src = source[name]
        rows = list(src.iter_rows(values_only=True))
        header_idx = _find_header_index(rows, _POS_HEADERS[name])
        banner = rows[header_idx - 1] if header_idx > 0 else tuple()
        header = rows[header_idx]
        data = rows[header_idx + 1 : header_idx + 1 + _DATA_ROWS]

        if name == "Deutsch":
            de_header_cols = {
                str(c).strip(): i for i, c in enumerate(header) if c is not None
            }

        dest = out.create_sheet(title=name)
        dest.append(list(banner))
        dest.append(list(header))
        for offset, row in enumerate(data):
            dest.append(_doctor(name, offset, list(row), de_header_cols))

    source.close()
    out.save(_FIXTURE)
    return _FIXTURE


def _doctor(sheet: str, offset: int, row: list, de_cols: dict[str, int]) -> list:
    """Apply the fixture-only edge-case doctoring to a single row."""

    if offset == _FR_IT_EMPTY_ROW and sheet in {"Français", "Italiano"}:
        # Blank the designation cell (column index 3 in every sheet).
        if len(row) > 3:
            row[3] = None
    if offset == _TP_TEXT_ROW and sheet == "Deutsch":
        # TP is column index 2; replace the number with non-billing free text.
        if len(row) > 2:
            row[2] = "nach Aufwand"
    if offset == _NO_DISCIPLINE_ROW and sheet == "Deutsch":
        for column in _DISCIPLINE_COLUMNS:
            i = de_cols.get(column)
            if i is not None and i < len(row):
                row[i] = None
    return row


if __name__ == "__main__":
    path = build()
    print(f"wrote fixture: {path}")
