"""BAG EAL adapter + pipeline tests — fully offline (SQLite mirror, no network).

The fixture ``analysenliste_2026-01-01_fixture.xlsx`` is built from the real workbook
by ``tests/fixtures/eal/make_fixture.py``: banner + header + 25 real parallel rows,
with three rows deliberately doctored to exercise edge paths the 2026-01-01 edition
does not itself contain (FR/IT-empty, non-numeric TP, no discipline flag).
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from tarifhub_ingest.adapters import bag_eal
from tarifhub_ingest.audit.audit_logger import AuditLogger
from tarifhub_ingest.config import get_settings
from tarifhub_ingest.ingestion.pipeline import run_pipeline
from tarifhub_ingest.ingestion.source_loader import SourceSpec
from tarifhub_ingest.mappers.tariff_mapper import map_raw
from tarifhub_ingest.models.tariff_model import TariffSystem
from tarifhub_ingest.storage.db import Database
from tarifhub_ingest.storage.tariff_repository import TariffRepository

_FIXTURE = (
    Path(__file__).resolve().parent / "fixtures" / "eal" / "analysenliste_2026-01-01_fixture.xlsx"
)
_EAL_URL = "https://www.bag.admin.ch/de/analysenliste-al"

# Pinned content hash of the frozen Pos-1000 record as produced by the pipeline
# (map_raw -> deterministic scorer confidence 1.0, requires_review False -> freeze).
# To regenerate: parse the fixture, map_raw Pos 1000 with the EAL system + source_url,
# set harmonization_confidence=score(rec) and requires_review=(score<0.85), freeze,
# and read frozen.record_hash. A change here means the hashing rule or the canonical
# Pos-1000 content changed — which must be deliberate.
POS_1000_HASH = "33f658ff57952693dc45b51c2c7b11e567d2d3799278247d3231aa67e40f69a0"


def _build_workbook(
    path: Path,
    *,
    de_headers: list[str],
    de_rows: list[list],
    fr_rows: list[list] | None = None,
    it_rows: list[list] | None = None,
) -> Path:
    """Write a minimal EAL-shaped workbook (banner + header + data) for guard tests.

    The DE sheet uses ``de_headers``; FR/IT sheets carry their own position-header
    label and a single 'Dénomination'/'Denominazione' column so duplicate-position
    guards can be exercised without depending on the large committed fixture.
    """
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    de = wb.create_sheet("Deutsch")
    de.append(["Fachbereiche"])  # banner row
    de.append(de_headers)
    for row in de_rows:
        de.append(row)

    if fr_rows is not None:
        fr = wb.create_sheet("Français")
        fr.append(["Domaines"])
        fr.append(["No.", "No. Pos.", "PT", "Dénomination"])
        for row in fr_rows:
            fr.append(row)

    if it_rows is not None:
        it = wb.create_sheet("Italiano")
        it.append(["Settori"])
        it.append(["No.", "No. pos.", "PT", "Denominazione"])
        for row in it_rows:
            it.append(row)

    wb.save(path)
    return path


# Canonical minimal DE header row (only the columns the adapter reads need be real).
_DE_HEADERS = ["Kapitel", "Pos.-Nr.", "TP", "Bezeichnung", "Chemie"]


@pytest.fixture(scope="module")
def rows() -> list[dict]:
    return bag_eal.parse(_FIXTURE)


def _by_code(rows: list[dict], code: str) -> dict:
    matches = [r for r in rows if r["tariff_code"] == code]
    assert matches, f"position {code} not in fixture"
    return matches[0]


def test_adapter_version_pinned():
    assert bag_eal.ADAPTER_VERSION == "bag-eal/0.1.0"


def test_golden_parse_pos_1000(rows):
    """Pos 1000 round-trips through the mapper with the expected canonical values."""
    raw = _by_code(rows, "1000")
    record = map_raw(raw, system=TariffSystem.EAL, source_url=_EAL_URL)

    assert record.tariff_code == "1000"
    assert record.tariff_system is TariffSystem.EAL
    assert record.designation.de == "1,25-Dihydroxy-Vitamin D"
    # FR + IT joins by position number are present and parallel.
    assert record.designation.fr == "1,25-dihydroxy-vitamine D"
    assert record.designation.it == "1,25-diidrossivitamina D"
    assert record.category == "Chemie"
    # Tax points are Decimal-typed (never float) and equal to 76.5.
    assert record.tax_points == Decimal("76.5")
    assert isinstance(record.tax_points, Decimal)
    assert record.price_chf is None
    assert record.unit == "point"
    assert record.valid_from == date(2026, 1, 1)
    assert record.source_version == "BAG AL 2026-01-01"


def test_edge_fr_it_empty_row(rows):
    """Doctored row 1012: FR + IT designation cells emptied -> None, not ''."""
    raw = _by_code(rows, "1012")
    assert raw["designation_fr"] is None
    assert raw["designation_it"] is None
    record = map_raw(raw, system=TariffSystem.EAL, source_url=_EAL_URL)
    assert record.designation.de  # German still present
    assert record.designation.fr is None
    assert record.designation.it is None


def test_edge_non_numeric_tp_row(rows):
    """Doctored row 1021: TP = 'nach Aufwand' -> tax_points None, unit None."""
    raw = _by_code(rows, "1021")
    assert raw["tax_points"] is None
    assert raw["unit"] is None
    record = map_raw(raw, system=TariffSystem.EAL, source_url=_EAL_URL)
    assert record.tax_points is None
    assert record.price_chf is None
    assert record.unit is None


def test_edge_no_discipline_row(rows):
    """Doctored row 1026: every discipline flag cleared -> category None."""
    raw = _by_code(rows, "1026")
    assert raw["category"] is None
    record = map_raw(raw, system=TariffSystem.EAL, source_url=_EAL_URL)
    assert record.category is None


def test_category_is_first_discipline_in_priority_order(rows):
    """A normal row resolves its category to the first 'Ja' discipline column."""
    raw = _by_code(rows, "1000")
    assert raw["category"] == "Chemie"  # first column flagged 'Ja'


def test_valid_from_derived_from_filename(tmp_path):
    """valid_from / source_version are a pure function of the filename convention."""
    assert bag_eal._date_from_filename("analysenliste_2026-07-01.xlsx") == date(2026, 7, 1)
    assert bag_eal._date_from_filename("analysenliste_2026-07-01.xlsx") is not None
    assert (
        bag_eal._source_version_from_filename("analysenliste_2026-07-01.xlsx")
        == "BAG AL 2026-07-01"
    )
    # A non-conforming name yields None (documented graceful degradation).
    assert bag_eal._date_from_filename("random.xlsx") is None
    assert bag_eal._source_version_from_filename("random.xlsx") is None


def test_rows_without_position_are_skipped(rows):
    """Every emitted row carries a non-empty tariff_code (no banner/blank rows)."""
    assert len(rows) == 25
    assert all(r["tariff_code"] for r in rows)


def test_refuses_oversized_file(tmp_path, monkeypatch):
    """Hostile-input guard: a file over the byte cap raises a clear ValueError."""
    big = tmp_path / "analysenliste_2026-01-01.xlsx"
    big.write_bytes(b"x")
    monkeypatch.setattr(bag_eal, "_MAX_BYTES", 0)
    with pytest.raises(ValueError, match="over the"):
        bag_eal.parse(big)


# --------------------------------------------------------------------------- #
# Ship-review fixes (1-4)
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize("dropped", ["TP", "Bezeichnung"])
def test_fail_fast_on_missing_required_de_header(tmp_path, dropped):
    """Fix 1: a renamed REQUIRED German column raises a ValueError naming the header.

    The position column ('Pos.-Nr.') also anchors the header row, so it is covered
    separately in :func:`test_fail_fast_on_renamed_position_header`.
    """
    headers = list(_DE_HEADERS)
    headers[headers.index(dropped)] = dropped + "_RENAMED"

    path = _build_workbook(
        tmp_path / "analysenliste_2026-01-01.xlsx",
        de_headers=headers,
        de_rows=[["k", "1000", 76.5, "Vitamin D", "Ja"]],
    )
    with pytest.raises(ValueError, match=dropped):
        bag_eal.parse(path)


def test_fail_fast_on_renamed_position_header(tmp_path):
    """Fix 1: if 'Pos.-Nr.' is renamed the anchor is lost; we raise rather than emit
    rows with silently-None designation/TP (the required-header check names it)."""
    headers = ["Kapitel", "Pos.-Nr._RENAMED", "TP", "Bezeichnung", "Chemie"]
    path = _build_workbook(
        tmp_path / "analysenliste_2026-01-01.xlsx",
        de_headers=headers,
        de_rows=[["k", "1000", 76.5, "Vitamin D", "Ja"]],
    )
    with pytest.raises(ValueError, match="Pos\\.-Nr\\."):
        bag_eal.parse(path)


def test_reject_duplicate_position_in_de_sheet(tmp_path):
    """Fix 2: a repeated Pos.-Nr. in the German sheet raises, naming the position."""
    path = _build_workbook(
        tmp_path / "analysenliste_2026-01-01.xlsx",
        de_headers=_DE_HEADERS,
        de_rows=[
            ["k", "1000", 76.5, "Vitamin D", "Ja"],
            ["k", "1000", 8.5, "Hämatokrit", "Ja"],  # duplicate join key
        ],
    )
    with pytest.raises(ValueError, match="1000"):
        bag_eal.parse(path)


def test_reject_duplicate_position_in_translation_sheet(tmp_path):
    """Fix 2: a repeated position in a translation sheet raises, naming the position."""
    path = _build_workbook(
        tmp_path / "analysenliste_2026-01-01.xlsx",
        de_headers=_DE_HEADERS,
        de_rows=[["k", "1000", 76.5, "Vitamin D", "Ja"]],
        fr_rows=[
            ["n", "1000", 76.5, "Vitamine D"],
            ["n", "1000", 76.5, "Vitamine D bis"],  # duplicate join key
        ],
    )
    with pytest.raises(ValueError, match="1000"):
        bag_eal.parse(path)


@pytest.mark.parametrize(
    "bad_url",
    [
        "http://www.bag.admin.ch/al.xlsx",  # wrong scheme
        "file:///etc/passwd",  # file://
        "https://evil.example.com/al.xlsx",  # wrong host
        "https://bag.admin.ch.evil.com/al.xlsx",  # suffix-spoof host
        "ftp://bag.admin.ch/al.xlsx",  # wrong scheme
    ],
)
def test_fetch_rejects_untrusted_url(bad_url):
    """Fix 3 (M1): fetch validates scheme=https + host bag.admin.ch up front (no network)."""
    with pytest.raises(ValueError):
        bag_eal.fetch(bad_url, "/tmp/should-never-be-written.xlsx")


@pytest.mark.parametrize(
    "good_url",
    [
        "https://bag.admin.ch/al.xlsx",
        "https://www.bag.admin.ch/de/analysenliste-al",
    ],
)
def test_fetch_url_validation_accepts_bag_domain(good_url):
    """Fix 3: the validator itself accepts the federal domain + subdomains."""
    # Pure-function check on the validator — no download is attempted.
    bag_eal._validate_fetch_url(good_url)


def test_fetch_refuses_redirect(tmp_path, monkeypatch):
    """The scheme/host pin only covers the initial URL, so a 30x is refused (no network).

    The shared redirect-refusing opener raises before any byte is read off an
    unpinned host; fetch must surface that as a ValueError naming the redirect."""
    import urllib.request

    def _fake_open(request, *, timeout):
        raise ValueError("fetch refused a 302 redirect to 'https://evil.example.com/x'")

    monkeypatch.setattr(urllib.request, "Request", lambda *a, **k: object())
    monkeypatch.setattr(bag_eal, "open_no_redirect", _fake_open)
    with pytest.raises(ValueError, match="redirect"):
        bag_eal.fetch("https://bag.admin.ch/al.xlsx", tmp_path / "al.xlsx")


def test_row_limit_is_exact_not_off_by_one(tmp_path, monkeypatch):
    """Fix 4 (M2): with _MAX_ROWS=1, two data rows must raise (not silently keep 2)."""
    monkeypatch.setattr(bag_eal, "_MAX_ROWS", 1)
    path = _build_workbook(
        tmp_path / "analysenliste_2026-01-01.xlsx",
        de_headers=_DE_HEADERS,
        de_rows=[
            ["k", "1000", 76.5, "Vitamin D", "Ja"],
            ["k", "1001", 8.5, "Hämatokrit", "Ja"],
        ],
    )
    with pytest.raises(ValueError, match="row limit"):
        bag_eal.parse(path)

    # And exactly _MAX_ROWS rows is accepted (boundary is inclusive of the limit).
    ok = _build_workbook(
        tmp_path / "analysenliste_2026-01-01_ok.xlsx",
        de_headers=_DE_HEADERS,
        de_rows=[["k", "1000", 76.5, "Vitamin D", "Ja"]],
    )
    assert len(bag_eal.parse(ok)) == 1


def test_parse_missing_file_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        bag_eal.parse(tmp_path / "nope.xlsx")


def _pipeline(tmp_path, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    db = Database.from_url(f"sqlite:///{tmp_path / 'eal.db'}")
    conn = db.connect()
    db.init_schema(conn)
    repo = TariffRepository(conn, db)
    audit = AuditLogger(conn, db)
    spec = SourceSpec(system=TariffSystem.EAL, kind="bag_eal", path=_FIXTURE, source_url=_EAL_URL)
    return [spec], repo, audit, conn


def test_pipeline_source_to_freeze(tmp_path, monkeypatch):
    """End-to-end source->freeze over the fixture on SQLite, kind 'bag_eal'."""
    specs, repo, audit, conn = _pipeline(tmp_path, monkeypatch)
    settings = get_settings()

    report = run_pipeline(specs, repo, audit, settings=settings)

    assert report.processed == 25
    assert report.frozen == 25
    assert report.skipped_existing == 0

    # The TP-less doctored row (1021) scores below 0.85 -> requires_review.
    stored = repo.get("1021", TariffSystem.EAL)
    assert stored is not None
    assert stored.tax_points is None
    assert stored.requires_review is True
    assert report.flagged_for_review >= 1

    # The complete golden row is high-confidence -> not flagged.
    golden = repo.get("1000", TariffSystem.EAL)
    assert golden is not None
    assert golden.requires_review is False
    conn.close()


def test_pipeline_pins_pos_1000_record_hash(tmp_path, monkeypatch):
    """Pinned-hash regression for one known fixture record (Pos 1000)."""
    specs, repo, audit, conn = _pipeline(tmp_path, monkeypatch)
    run_pipeline(specs, repo, audit, settings=get_settings())

    golden = repo.get("1000", TariffSystem.EAL)
    assert golden is not None
    assert golden.record_hash == POS_1000_HASH
    conn.close()


def test_pipeline_is_idempotent_on_rerun(tmp_path, monkeypatch):
    """A second run over the same fixture skips every record (idempotency)."""
    specs, repo, audit, conn = _pipeline(tmp_path, monkeypatch)
    settings = get_settings()

    first = run_pipeline(specs, repo, audit, settings=settings)
    second = run_pipeline(specs, repo, audit, settings=settings)

    assert first.frozen == 25
    assert second.frozen == 0
    assert second.skipped_existing == first.frozen
    conn.close()
