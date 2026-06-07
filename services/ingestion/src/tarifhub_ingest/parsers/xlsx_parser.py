"""XLSX / CSV parser for flat tabular tariff sources (e.g. BAG EAL lab analyses).

Emits a list of flat ``dict`` rows keyed by the source header names. Values keep
their native types (str/number/date); the mapper coerces them. No AI, no network.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

PARSER_VERSION = "xlsx-parser/0.1.0"


def parse(path: str | Path) -> list[dict[str, Any]]:
    """Parse an ``.xlsx`` or ``.csv`` file into a list of row dicts."""

    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(path)
    if suffix == ".csv":
        return parse_csv(path)
    raise ValueError(f"unsupported tabular source: {path.name}")


def parse_xlsx(path: str | Path) -> list[dict[str, Any]]:
    """Parse the first worksheet of an XLSX workbook (header row + data rows)."""

    # Imported lazily so a CSV-only run does not require openpyxl to be present.
    from openpyxl import load_workbook  # noqa: PLC0415

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = _clean_headers(next(rows, None))
        if not headers:
            return []
        records: list[dict[str, Any]] = []
        for row in rows:
            record = _row_to_dict(headers, row)
            if record:
                records.append(record)
        return records
    finally:
        workbook.close()


def parse_csv(path: str | Path) -> list[dict[str, Any]]:
    """Parse a CSV file (UTF-8, header row + data rows)."""

    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        headers = _clean_headers(next(reader, None))
        if not headers:
            return []
        records: list[dict[str, Any]] = []
        for row in reader:
            record = _row_to_dict(headers, row)
            if record:
                records.append(record)
        return records


def _clean_headers(row: Any) -> list[str]:
    if not row:
        return []
    return [str(cell).strip() if cell is not None else "" for cell in row]


def _row_to_dict(headers: list[str], row: Any) -> dict[str, Any]:
    if row is None:
        return {}
    values = list(row)
    record: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = values[index] if index < len(values) else None
        if isinstance(value, str):
            value = value.strip()
        record[header] = value
    # Drop rows where every mapped value is empty.
    if all(v is None or v == "" for v in record.values()):
        return {}
    return record
