"""Adapter for the BAG Analysenliste (EAL) — the real Swiss lab-analysis tariff.

Source: https://www.bag.admin.ch/de/analysenliste-al ("Analysenliste per …",
Excel format). One workbook, three parallel sheets ``Deutsch`` / ``Français`` /
``Italiano`` (1279 data rows each), joined by position number.

``parse`` is a pure function of the file (no AI, no network): it emits a list of
canonical-keyed ``dict`` rows that :func:`tarifhub_ingest.mappers.tariff_mapper.map_raw`
consumes. ``fetch`` (stdlib only, used by the scale-run driver, never by tests)
downloads the artifact and writes a ``.sha256`` sidecar idempotently.

``valid_from`` / ``source_version`` are derived purely from the FILENAME convention
``analysenliste_YYYY-MM-DD.xlsx`` — a deterministic function of the input path, so
re-ingesting the same file always yields the same frozen records.
"""

from __future__ import annotations

import hashlib
import re
import urllib.request
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

ADAPTER_VERSION = "bag-eal/0.1.0"

# Hostile-input guards: the real file is ~0.8 MB / 1281 rows; refuse anything wildly
# outside that envelope rather than letting a malformed/oversized file exhaust memory.
_MAX_BYTES = 20 * 1024 * 1024
_MAX_ROWS = 100_000

# Network guard for fetch(): stream-cap the download at the same byte budget, and
# pin the scheme/host so a future caller passing untrusted input can't reach
# file:// / internal addresses (SSRF). Only https on the BAG federal domain.
_FETCH_TIMEOUT_S = 30
_FETCH_MAX_BYTES = 20 * 1024 * 1024
_CHUNK = 64 * 1024
_ALLOWED_SCHEME = "https"
_ALLOWED_HOST = "bag.admin.ch"

# Header anchor: the row whose SECOND cell (stripped) equals this is the header row.
# We search the first few rows rather than hardcoding an index, so a future edition
# that adds/removes a banner row above the header still parses.
_POS_HEADER_DE = "Pos.-Nr."
_POS_HEADER_FR = "No. Pos."
_POS_HEADER_IT = "No. pos."
_HEADER_SEARCH_ROWS = 10

# Column headers (stripped) within each language sheet.
_H_POS_DE = "Pos.-Nr."
_H_TP_DE = "TP"
_H_DESIG_DE = "Bezeichnung"
_H_POS_FR = "No. Pos."
_H_DESIG_FR = "Dénomination"
_H_POS_IT = "No. pos."
_H_DESIG_IT = "Denominazione"

# Discipline flag columns in canonical priority order; the first column carrying the
# exact value "Ja" is taken as the record's category.
_DISCIPLINE_COLUMNS = (
    "Chemie",
    "Hämatologie",
    "Immunologie",
    "Genetik",
    "Mikrobiologie",
    "Spezialanalyse",
    "Basisanalyse",
)
_DISCIPLINE_FLAG = "Ja"

# analysenliste_YYYY-MM-DD.xlsx  ->  the publication date.
_FILENAME_DATE_RE = re.compile(r"analysenliste_(\d{4})-(\d{2})-(\d{2})", re.IGNORECASE)


def parse(path: str | Path) -> list[dict[str, Any]]:
    """Parse a BAG EAL workbook into canonical-keyed row dicts.

    Emits one dict per German data row (the canonical language). French/Italian
    designations are joined by position number; a missing FR/IT sheet or a missing
    counterpart row degrades gracefully to a German-only row. Rows without a
    position number are skipped (they cannot be a tariff position).
    """

    path = Path(path)
    _guard_file_size(path)

    from openpyxl import load_workbook  # noqa: PLC0415  (lazy: keep CSV-only runs light)

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheets = {s.title: s for s in workbook.worksheets}
        de_sheet = _pick_sheet(sheets, ("Deutsch",), _POS_HEADER_DE)
        if de_sheet is None:
            raise ValueError("BAG EAL workbook has no German ('Deutsch') sheet")

        # Fail fast if BAG renamed a REQUIRED German column: silently emitting rows
        # with None designation/TP would freeze junk. (FR/IT stay tolerant by design.)
        de_headers = _find_header_row(de_sheet, _POS_HEADER_DE) or []
        _require_headers(de_headers, (_H_POS_DE, _H_TP_DE, _H_DESIG_DE))

        fr_by_pos = _index_translations(
            _pick_sheet(sheets, ("Français", "Francais"), _POS_HEADER_FR),
            _POS_HEADER_FR,
            _H_POS_FR,
            _H_DESIG_FR,
            sheet_label="Français",
        )
        it_by_pos = _index_translations(
            _pick_sheet(sheets, ("Italiano",), _POS_HEADER_IT),
            _POS_HEADER_IT,
            _H_POS_IT,
            _H_DESIG_IT,
            sheet_label="Italiano",
        )

        valid_from = _date_from_filename(path.name)
        source_version = _source_version_from_filename(path.name)

        records: list[dict[str, Any]] = []
        seen_positions: set[str] = set()
        for headers, values in _data_rows(de_sheet, _POS_HEADER_DE):
            index = {h: i for i, h in enumerate(headers)}
            pos = _cell(values, index.get(_H_POS_DE))
            tariff_code = _as_str(pos)
            if not tariff_code:
                continue  # no position number -> not a tariff position

            # The position number is the frozen join key; a duplicate would silently
            # overwrite translations / collide on the (system, code) version key.
            if tariff_code in seen_positions:
                raise ValueError(
                    f"BAG EAL 'Deutsch' sheet repeats position {tariff_code!r}; "
                    "position numbers must be unique (frozen join key)"
                )
            seen_positions.add(tariff_code)

            if len(records) >= _MAX_ROWS:
                raise ValueError(
                    f"BAG EAL workbook exceeds the {_MAX_ROWS} row limit; refusing to parse"
                )
            records.append(
                {
                    "tariff_code": tariff_code,
                    "designation_de": _as_str(_cell(values, index.get(_H_DESIG_DE))),
                    "designation_fr": fr_by_pos.get(tariff_code),
                    "designation_it": it_by_pos.get(tariff_code),
                    "category": _category(values, index),
                    "tax_points": _tax_points(_cell(values, index.get(_H_TP_DE))),
                    "unit": _unit(_cell(values, index.get(_H_TP_DE))),
                    "valid_from": valid_from,
                    "source_version": source_version,
                }
            )
        return records
    finally:
        workbook.close()


def fetch(url: str, dest_path: str | Path) -> Path:
    """Download ``url`` to ``dest_path`` (stdlib only) and write a ``.sha256`` sidecar.

    Idempotent: if the destination already exists and its current hash matches the
    sidecar, the download is skipped. NOT used by the test suite (no network in
    tests) — this is the scale-run driver's entry point.
    """

    _validate_fetch_url(url)
    dest = Path(dest_path)
    sidecar = dest.with_name(dest.name + ".sha256")

    if dest.exists() and sidecar.exists():
        recorded = sidecar.read_text(encoding="utf-8").split()[0].strip()
        if recorded and recorded == _sha256_file(dest):
            return dest  # already present and intact

    dest.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "tarifhub-ingest"})
    digest = hashlib.sha256()
    total = 0
    with urllib.request.urlopen(  # noqa: S310 — caller-supplied source URL, http(s)
        request, timeout=_FETCH_TIMEOUT_S
    ) as response, open(dest, "wb") as handle:
        while True:
            chunk = response.read(_CHUNK)
            if not chunk:
                break
            total += len(chunk)
            if total > _FETCH_MAX_BYTES:
                handle.close()
                dest.unlink(missing_ok=True)
                raise ValueError(
                    f"download exceeded {_FETCH_MAX_BYTES} bytes; aborted before completion"
                )
            handle.write(chunk)
            digest.update(chunk)

    sidecar.write_text(f"{digest.hexdigest()}  {dest.name}\n", encoding="utf-8")
    return dest


# --------------------------------------------------------------------------- #
# Internals
# --------------------------------------------------------------------------- #


def _validate_fetch_url(url: str) -> None:
    """Reject anything but https on the BAG federal domain (SSRF / file:// guard)."""

    parts = urlsplit(url)
    if parts.scheme != _ALLOWED_SCHEME:
        raise ValueError(
            f"fetch URL scheme must be {_ALLOWED_SCHEME!r}, got {parts.scheme!r}: {url}"
        )
    host = (parts.hostname or "").lower()
    if host != _ALLOWED_HOST and not host.endswith("." + _ALLOWED_HOST):
        raise ValueError(
            f"fetch URL host must be {_ALLOWED_HOST!r} or a subdomain, got {host!r}: {url}"
        )


def _require_headers(headers: list[str], required: tuple[str, ...]) -> None:
    """Raise if any required German column header is absent after anchoring."""

    present = set(headers)
    missing = [h for h in required if h not in present]
    if missing:
        raise ValueError(
            "BAG EAL 'Deutsch' sheet is missing required column header(s): "
            + ", ".join(repr(h) for h in missing)
        )


def _guard_file_size(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"BAG EAL source not found: {path}")
    size = path.stat().st_size
    if size > _MAX_BYTES:
        raise ValueError(
            f"BAG EAL source is {size} bytes, over the {_MAX_BYTES} limit; refusing to parse"
        )


def _pick_sheet(sheets: dict[str, Any], names: tuple[str, ...], pos_header: str) -> Any | None:
    """Return the sheet by preferred title, falling back to a header-anchor scan."""

    for name in names:
        if name in sheets:
            return sheets[name]
    # Fall back: any sheet whose header row carries the expected position header.
    for sheet in sheets.values():
        if _find_header_row(sheet, pos_header) is not None:
            return sheet
    return None


def _find_header_row(sheet: Any, pos_header: str) -> list[str] | None:
    """Find the header row by locating the row whose 2nd cell equals ``pos_header``.

    Robust to a banner row above the header: scans the first ``_HEADER_SEARCH_ROWS``
    rows rather than assuming a fixed index.
    """

    for count, row in enumerate(sheet.iter_rows(values_only=True)):
        if count >= _HEADER_SEARCH_ROWS:
            break
        cells = list(row)
        if len(cells) > 1 and _as_str(cells[1]) == pos_header:
            return [_as_str(c) or "" for c in cells]
    return None


def _data_rows(sheet: Any, pos_header: str):
    """Yield ``(headers, values)`` for each row beneath the located header row."""

    headers: list[str] | None = None
    header_seen = False
    for count, row in enumerate(sheet.iter_rows(values_only=True)):
        cells = list(row)
        if not header_seen:
            if count < _HEADER_SEARCH_ROWS and len(cells) > 1 and _as_str(cells[1]) == pos_header:
                headers = [_as_str(c) or "" for c in cells]
                header_seen = True
            continue
        if headers is not None:
            yield headers, cells


def _index_translations(
    sheet: Any | None,
    pos_header: str,
    pos_col: str,
    desig_col: str,
    *,
    sheet_label: str,
) -> dict[str, str | None]:
    """Build ``{position -> designation}`` for a translation sheet (empty if absent).

    A missing sheet/column is tolerated (de-only rows, by design), but a *repeated*
    position number is rejected: it is the frozen join key, and a silent overwrite of
    one translation by another is unacceptable.
    """

    index: dict[str, str | None] = {}
    if sheet is None:
        return index
    for headers, values in _data_rows(sheet, pos_header):
        col = {h: i for i, h in enumerate(headers)}
        pos = _as_str(_cell(values, col.get(pos_col)))
        if not pos:
            continue
        if pos in index:
            raise ValueError(
                f"BAG EAL {sheet_label!r} sheet repeats position {pos!r}; "
                "position numbers must be unique (frozen join key)"
            )
        index[pos] = _as_str(_cell(values, col.get(desig_col)))
    return index


def _category(values: list[Any], index: dict[str, int]) -> str | None:
    """First discipline column (in canonical order) flagged exactly 'Ja', else None."""

    for column in _DISCIPLINE_COLUMNS:
        i = index.get(column)
        if i is not None and _as_str(_cell(values, i)) == _DISCIPLINE_FLAG:
            return column
    return None


def _tax_points(value: Any) -> Any:
    """Return the TP value only when numeric; non-numeric ('nach Aufwand') -> None.

    The mapper coerces this to ``Decimal``; we pass numbers straight through and
    drop anything non-numeric so the canonical record carries no junk billing value.
    """

    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = _as_str(value)
    if not text:
        return None
    try:
        float(text.replace("'", "").replace(",", "."))
    except ValueError:
        return None
    return value


def _unit(tp_value: Any) -> str | None:
    """EAL is tax-point based: unit is 'point' iff a numeric TP is present."""

    return "point" if _tax_points(tp_value) is not None else None


def _date_from_filename(name: str) -> date | None:
    match = _FILENAME_DATE_RE.search(name)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _source_version_from_filename(name: str) -> str | None:
    parsed = _date_from_filename(name)
    return f"BAG AL {parsed.isoformat()}" if parsed else None


def _cell(values: list[Any], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(_CHUNK), b""):
            digest.update(chunk)
    return digest.hexdigest()
